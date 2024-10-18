from datetime import datetime
import datetime
import openpyxl
import pandas as pd
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from .models import Megacamp
from .serializers import MegacampSerializer
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Count

class MegacampAPI(ModelViewSet):
    queryset = Megacamp.objects.all()
    serializer_class = MegacampSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            api_response = {
                'status': 'success',
                'code': status.HTTP_201_CREATED,
                'message': 'Mega Camp record added successfully',
                'new_record': serializer.data,
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to add MegaCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.serializer_class(instance)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'megacamp': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve MegaCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_404_NOT_FOUND,
                'message': error_message
            }
            return Response(error_response)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.serializer_class(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'MegaCamp record updated successfully',
                'updated_record': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to update MegaCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def partial_update(self, request, *args, **kwargs):
        try:
            kwargs['partial'] = True
            response = self.update(request, *args, **kwargs)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'Mega Camp record updated successfully',
                'updated_record': response.data
            }
            return Response(api_response, status=status.HTTP_200_OK)
        except Exception as e:
            error_message = 'Failed to partially update MegaCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            instance.delete()
            api_response = {
                'status': 'success',
                'code': status.HTTP_204_NO_CONTENT,
                'message': 'Mega Camp record deleted successfully'
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to delete MegaCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.serializer_class(queryset, many=True)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'megacamps': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve MegaCamp records: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)


class MegacampExcelSheet(APIView):
    def get(self, request):
        data = Megacamp.objects.all()
        headers_mapping = {
            'srno' :' Sr No',
            'village' : 'Main Village',
            'date': 'Date',
            'year': 'Year',
            'month': 'Month',
            'day':'Day',
            'name' : 'Name',
            'gender': 'Gender',
            'age' : 'Age',
            'contact': 'Contact',
            'villagename' : 'Village Name',
            'weight': 'Weight',
            'height': 'Height',
            'bp':'BP',
            'pulse': 'Pulse',
            'temperature':'Temperature',
            'bloodtest':'Blood Test',
            'hb': 'HB',
            'xray': 'X-Ray',
            'ecg': 'ECG',
            'eyetest': 'Eye Test',
            'audiometry':'Audiometry',
            'spirometry':'Spirometry',
            'breastcancer':'Breast Cancer',
            'cervicalcancer': 'Cervical Cancer',
            'oralcancer': 'Oral Cancer',
            'tb': 'TB',
            'description': 'Description',
            'client_name': 'Client Name'
        }
        wb = openpyxl.Workbook()
        ws=wb.active
        ws.merge_cells('A1:AC1')
        ws['A1'] = 'Mega Camp Report'
        ws['A1'].font = Font(bold= True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers= list(headers_mapping.values())
        ws.append(headers)

        for cell in ws[2]:
            cell.font = Font(bold=True)
        for row in data:
            row_data=[getattr(row,field) for field in headers_mapping.keys()]
            ws.append(row_data)
            # Bold the header row

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=29):
            for cell in row:
                cell.border = thin_border

        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Mega CAMP {current_date}.xlsx"'
        wb.save(response)
        return response



