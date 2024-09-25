from datetime import datetime
import datetime
import openpyxl
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from .models import EyeCamp
from .serializers import EyeCampSerializer
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class EyeCampAPI(ModelViewSet):
    queryset = EyeCamp.objects.all()
    serializer_class = EyeCampSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            api_response = {
                'status': 'success',
                'code': status.HTTP_201_CREATED,
                'message': 'EyeCamp record added successfully',
                'new_record': serializer.data,
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to add EyeCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.serializer_class(instance)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'Eyecamp': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve EyeCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_404_NOT_FOUND,
                'message': error_message
            }
            return Response(error_response)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.serializer_class(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'EyeCamp record updated successfully',
                'updated_record': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to update EyeCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def partial_update(self, request, *args, **kwargs):
        try:
            kwargs['partial'] = True
            response = self.update(request, *args, **kwargs)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'EyeCamp record updated successfully',
                'updated_record': response.data
            }
            return Response(api_response, status=status.HTTP_200_OK)
        except Exception as e:
            error_message = 'Failed to partially update EyeCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            instance.delete()
            api_response = {
                'status': 'success',
                'code': status.HTTP_204_NO_CONTENT,
                'message': 'EyeCamp record deleted successfully'
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to delete EyeCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.serializer_class(queryset, many=True)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'Eyecamps': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve EyeCamp records: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)


class EyeCampExcelSheet(APIView):
    def get(self, request):
        data = EyeCamp.objects.all()
        headers_mapping = {
            'SrNo':' Sr No',
            'village': 'Main Village',
            'date': 'Date',
            'year': 'Year',
            'month': 'Month',
            'name': 'Name',
            'gender': 'Gender',
            'contact': 'Contact',
            'subvillage': 'Village Name',
            'age': 'Age',
            'code': 'Unique Code',
            'desciption': 'Code Description',
            'opinion': 'Opinion'

        }
        wb = openpyxl.Workbook()
        ws=wb.active
        ws.merge_cells('A1:L1')
        ws['A1'] = 'NCD- Village Level Eye Screening Camp'
        ws['A1'].font = Font(bold= True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers= list(headers_mapping.values())
        ws.append(headers)

        for cell in ws[2]:
            cell.font = Font(bold=True)
        for row in data:
            row_data=[getattr(row,field) for field in headers_mapping.keys()]
            ws.append(row_data)
            # Bold the header row

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
            for cell in row:
                cell.border = thin_border

        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Eye CAMP {current_date}.xlsx"'
        wb.save(response)
        return response

class EyeCampReport(APIView):
    def get(self,request):
        data = EyeCamp.objects.all()

