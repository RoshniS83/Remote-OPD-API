from datetime import datetime
import datetime
import openpyxl
import pandas as pd
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from .models import HBCamp
from .serializers import HBCampSerializer
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Count

class HBCampAPI(ModelViewSet):
    queryset = HBCamp.objects.all()
    serializer_class = HBCampSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            api_response = {
                'status': 'success',
                'code': status.HTTP_201_CREATED,
                'message': 'HBCamp record added successfully',
                'new_record': serializer.data,
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to add HBCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.serializer_class(instance)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'hbcamp': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve HBCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_404_NOT_FOUND,
                'message': error_message
            }
            return Response(error_response)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.serializer_class(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'HBCamp record updated successfully',
                'updated_record': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to update HBCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def partial_update(self, request, *args, **kwargs):
        try:
            kwargs['partial'] = True
            response = self.update(request, *args, **kwargs)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'HBCamp record updated successfully',
                'updated_record': response.data
            }
            return Response(api_response, status=status.HTTP_200_OK)
        except Exception as e:
            error_message = 'Failed to partially update HBCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            instance.delete()
            api_response = {
                'status': 'success',
                'code': status.HTTP_204_NO_CONTENT,
                'message': 'HBCamp record deleted successfully'
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to delete HBCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.serializer_class(queryset, many=True)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'hbcamps': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve HBCamp records: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)


class HBCampExcelSheet(APIView):
    def get(self, request):
        data = HBCamp.objects.all()
        headers_mapping = {
            'SrNo' :' Sr No',
            'village' : 'Main Village',
            'date': 'Date',
            'year': 'Year',
            'month': 'Month',
            'name' : 'Name',
            'gender' : 'Gender',
            'contact': 'Contact',
            'subvillage' : 'Village Name',
            'age' : 'Age',
            'HB' : 'HB',
            'HBReadings' : 'HBReadings'
        }
        wb = openpyxl.Workbook()
        ws=wb.active
        ws.merge_cells('A1:L1')
        ws['A1'] = 'NCD- Village Level HB Screening Camp'
        ws['A1'].font = Font(bold= True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers= list(headers_mapping.values())
        ws.append(headers)

        for cell in ws[2]:
            cell.font = Font(bold=True)
        for row in data:
            row_data=[getattr(row,field) for field in headers_mapping.keys()]
            ws.append(row_data)
            # Bold the header row

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
            for cell in row:
                cell.border = thin_border

        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="HB CAMP {current_date}.xlsx"'
        wb.save(response)
        return response


class HBCampMonthlyReport(APIView):
    def get(self, request):
        # Query data from the HBCamp model
        # data = HBCamp.objects.values('subvillage', 'gender', 'HBReadings').annotate(count=Count('SrNo'))
        # Get query parameters for filtering
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        village = request.query_params.get('village')

        # Validate required parameters
        if not year:
            return Response({"error": "Year is required"}, status=400)
        if not month:
            return Response({"error": "Month is required"}, status=400)
        if not village:
            return Response({"error": "Village is required"}, status=400)

            # Filter data from HBCamp model based on the provided parameters
        data = HBCamp.objects.filter(village=village, year=year, month=month).values('subvillage', 'gender',
                                                                                     'HBReadings').annotate(
            count=Count('SrNo'))

        # Convert to pandas dataframe
        df = pd.DataFrame(list(data))

        # Check if the DataFrame is empty
        if df.empty:
            return Response({"error": f"No data available for year: {year}, Month: {month}, Village: {village}"},
                            status=400)




            # Create a pivot table with 'subvillage' and 'gender' as index and 'HBReadings' as columns
        pivot_table = pd.pivot_table(df,
                                     index=['subvillage', 'gender'],
                                     columns='HBReadings',
                                     values='count',
                                     aggfunc='sum',
                                     fill_value=0)

        # Add Total column for each row
        pivot_table['Total'] = pivot_table.sum(axis=1)

        # Add Grand Total for each village by summing across genders
        grand_total = pivot_table.groupby('subvillage')['Total'].sum()
        pivot_table['Grand Total Village'] = pivot_table.index.get_level_values('subvillage').map(grand_total)

        # Reorder the columns to match the required order
        required_columns = ['Mild', 'Moderate', 'Severe', 'Healthy', 'Total', 'Grand Total Village']
        for col in required_columns:
            if col not in pivot_table.columns:
                pivot_table[col] = 0  # Add missing columns if necessary

        pivot_table = pivot_table[required_columns]

        # Convert pivot table to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'HB Camp Report'

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = 'Village Wise HB Screening Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Sr. No', 'Village Names', 'Gender'] + required_columns
        ws.append(headers)

        # Bold the header row
        for cell in ws[2]:
            cell.font = Font(bold=True)

            # Add data to the Excel sheet with merged village names
        sr_no = 1
        current_row = 3
        last_column_idx=9
        for village, group_data in pivot_table.groupby(level=0):
            # Number of rows to merge (male and female rows)
            num_rows = len(group_data)

            # Loop through each gender row in group_data (Male/Female)
            for idx, ((_, gender), row_data) in enumerate(group_data.iterrows()):
                ws.cell(row=current_row, column=1).value = sr_no  # Sr. No (we will merge this later)
                ws.cell(row=current_row, column=2).value = village  # Village Name (we will merge this later)
                ws.cell(row=current_row, column=3).value = gender  # Gender

                # Fill the rest of the values (Mild, Moderate, etc.)
                for col_idx, value in enumerate(row_data.tolist(), start=4):
                    ws.cell(row=current_row, column=col_idx).value = value  # Assign values for Mild, Moderate, etc.

                current_row += 1

            # Calculate Grand Total after looping over genders
            grand_total_value = group_data['Total'].sum()
            for r in range(current_row - num_rows, current_row):
                ws.cell(row=r, column=last_column_idx).value = grand_total_value  # Assign the Grand Total for each row

            # Now merge the Sr. No, Village, and Grand Total cells after writing the values
            ws.merge_cells(start_row=current_row - num_rows, start_column=1, end_row=current_row - 1,
                           end_column=1)  # Merge Sr. No
            ws.merge_cells(start_row=current_row - num_rows, start_column=2, end_row=current_row - 1,
                           end_column=2)  # Merge Village Name
            ws.merge_cells(start_row=current_row - num_rows, start_column=last_column_idx, end_row=current_row - 1,
                           end_column=last_column_idx)  # Merge Grand Total

            # Align merged cells
            ws.cell(row=current_row - num_rows, column=1).alignment = Alignment(vertical='center')
            ws.cell(row=current_row - num_rows, column=2).alignment = Alignment(vertical='center')
            ws.cell(row=current_row - num_rows, column=last_column_idx).alignment = Alignment(vertical='center')

            # Increment Sr. No for next village
            sr_no += 1

            # Apply thin border to all cells
        thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.border = thin_border

            # Prepare the response to download the file
        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="HB_CAMP_Report_{current_date}.xlsx"'
        wb.save(response)
        return response