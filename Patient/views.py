import io
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
import openpyxl
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import datetime
from Patient.models import Patientopdform
from Patient.serializers import PatientSerializer, PatientHistorySerializer
import pandas as pd
from django.db.models import Q
import os
from django.conf import settings
from django.http import FileResponse
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

MONTH_MAP = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6, 'July': 7, 'August': 8,
             'September': 9, 'October': 10, 'November': 11, 'December': 12}


# Create your views here.
class PatientAPI(ModelViewSet):
	queryset = Patientopdform.objects.all()
	serializer_class = PatientSerializer

	def create(self, request, *args, **kwargs):
		try:
			serializer = self.serializer_class(data=request.data)
			serializer.is_valid(raise_exception=True)
			serializer.save()
			api_response = {'status': 'success',
			                'code': status.HTTP_201_CREATED,
			                'message': 'Patient OPD form added successfully',
			                'new_form': serializer.data,
			                }
			return Response(api_response)
		except Exception as e:
			error_message = 'Failed to add patient opd form: {}'.format(str(e))
			error_response = {
				'status': 'error',
				'code': status.HTTP_400_BAD_REQUEST,
				'message': error_message
			}
			return Response(error_response)

	def retrieve(self, request, *args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.serializer_class(instance)
			api_response = {'status': 'success',
			                'code': status.HTTP_200_OK,
			                'patient': serializer.data}
			return Response(api_response)
		except Exception as e:
			error_message = 'Failed to retrieve patient opd form: {}'.format(str(e))
			error_response = {
				'status': 'error',
				'code': status.HTTP_404_NOT_FOUND,
				'message': error_message
			}
			return Response(error_response)

	def update(self, request, *args, **kwargs):
		try:
			partial = kwargs.pop('partial', False)
			instance = self.get_object()
			serializer = self.serializer_class(instance, data=request.data, partial=partial)
			serializer.is_valid(raise_exception=True)
			serializer.save()

			api_response = {
				'status': 'success',
				'code': status.HTTP_200_OK,
				'message': 'Patient OPD form updated successfully',
				'updated_form': serializer.data
			}
			return Response(api_response)
		except Exception as e:
			error_message = 'Failed to update patient opd form:{}'.format(str(e))
			error_response = {
				'status': 'error',
				'code': status.HTTP_400_BAD_REQUEST,
				'message': error_message
			}
			return Response(error_response)

	def partial_update(self, request, *args, **kwargs):
		try:
			kwargs['partial'] = True
			response = self.update(request, *args, **kwargs)
			api_response = {
				'status': 'success',
				'code': status.HTTP_200_OK,
				'message': 'Patient OPD form updated successfully',
				'updated_form': response.data
			}
			return Response(api_response, status=status.HTTP_200_OK)
		except Exception as e:
			error_message = f'Failed to partially update patient opd form: {str(e)} '
			error_response = {
				'status': 'error',
				'code': status.HTTP_400_BAD_REQUEST,
				'message': error_message
			}
			return Response(error_response)

	def destroy(self, request, *args, **kwargs):
		try:
			instance = self.get_object()
			instance.delete()
			api_response = {
				'status': 'success',
				'code': status.HTTP_204_NO_CONTENT,
				'message': 'Patient OPD form deleted successfully'
			}
			return Response(api_response)
		except Exception as e:
			error_message = 'Failed to delete patient opd form {}'.format(str(e))
			error_response = {
				'status': 'error',
				'code': status.HTTP_400_BAD_REQUEST,
				'message': error_message
			}
			return Response(error_response)

	# def list(self, request, *args, **kwargs):
	#           try:
	#                     # Extract the start index from the query parameters (default to 0 if not provided)
	#                     start_index = int(request.query_params.get('start_index', 0))
	#                     limit = 50  # Number of records to return per page
	#
	#                     # Get the queryset and ensure it's ordered for consistent pagination
	#                     queryset = self.get_queryset().order_by('pk')
	#
	#                     # Set up the Paginator
	#                     paginator = Paginator(queryset, limit)
	#                     page_number = (
	#                                             start_index // limit) + 1  # Calculate the page number based on the start index
	#
	#                     try:
	#                               page = paginator.page(page_number)
	#                     except PageNotAnInteger:
	#                               page = paginator.page(
	#                                         1)  # Default to the first page if page_number is not an integer
	#                     except EmptyPage:
	#                               return Response({"message": "No more records available"},
	#                                               status=status.HTTP_200_OK)
	#
	#                     # Serialize the paginated data
	#                     paginated_data = self.serializer_class(page.object_list, many=True)
	#
	#                     # Create the response object with pagination details
	#                     api_response = {
	#                               'status': 'success',
	#                               'code': status.HTTP_200_OK,
	#                               'total_records': paginator.count,  # Total number of records
	#                               'start_index': start_index,
	#                               'limit': limit,
	#                               'message': f'Patients starting from index {start_index}',
	#                               'patients': paginated_data.data,
	#                     }
	#
	#                     return Response(api_response, status=status.HTTP_200_OK)
	#
	#           except Exception as e:
	#                     error_message = f'Failed to retrieve patient opd forms: {str(e)}'
	#                     error_response = {
	#                               'status': 'error',
	#                               'code': status.HTTP_400_BAD_REQUEST,
	#                               'message': error_message
	#                     }
	#                     return Response(error_response, status=status.HTTP_400_BAD_REQUEST)
	def list(self, request, *args, **kwargs):
		try:
			# Get the month names from the query parameters
			from_month = request.query_params.get('from_month', 'January')  # Default to January
			to_month = request.query_params.get('to_month', 'December')  # Default to December
			year = int(request.query_params.get('year',datetime.datetime.now().year))  # Default to current year if not provided

			# Get the corresponding month numbers using the dictionary
			from_month_num = MONTH_MAP.get(from_month)
			to_month_num = MONTH_MAP.get(to_month)

			# get the search term from query parameters
			# search_term = request.query_params.get('search_term')

			# Extract the start index from the query parameters (default to 0 if not provided)
			start_index = int(request.query_params.get('start_index', 0))
			limit = 50  # Number of records to return per page

			# Filter the queryset based on the year and month range
			queryset = self.get_queryset().filter(
				date__year=year,
				date__month__gte=from_month_num,  # From selected month
				date__month__lte=to_month_num  # To selected month
			).order_by('pk')

			# Set up the Paginator
			paginator = Paginator(queryset, limit)
			page_number = (start_index // limit) + 1  # Calculate the page number based on the start index

			try:
				page = paginator.page(page_number)
			except PageNotAnInteger:
				page = paginator.page(1)  # Default to the first page if page_number is not an integer
			except EmptyPage:
				return Response({"message": "No more records available"}, status=status.HTTP_200_OK)

			# Serialize the paginated data
			paginated_data = self.serializer_class(page.object_list, many=True)

			# Create the response object with pagination details
			api_response = {
				'status': 'success',
				'code': status.HTTP_200_OK,
				'total_records': paginator.count,  # Total number of records
				'start_index': start_index,
				'limit': limit,
				'message': f'Patients from {from_month} to {to_month} in {year}',
				'patients': paginated_data.data,
			}

			return Response(api_response, status=status.HTTP_200_OK)

		except Exception as e:
			error_message = f'Failed to retrieve patient opd forms: {str(e)}'
			error_response = {
				'status': 'error',
				'code': status.HTTP_400_BAD_REQUEST,
				'message': error_message
			}
			return Response(error_response, status=status.HTTP_400_BAD_REQUEST)

	def search(self, request, *args, **kwargs):
		try:
			# Get the search term from query parameters
			search_term = request.query_params.get('search_term')
			if not search_term:
				return Response({"message": "Please provide a search term"},
				                status=status.HTTP_400_BAD_REQUEST)

			# Get the filters for year, month range, and village from the request query parameters
			from_month = request.query_params.get('from_month', 'January')
			to_month = request.query_params.get('to_month', 'December')
			year = int(request.query_params.get('year', datetime.datetime.now().year))

			# Map month names to numbers (assuming MONTH_MAP exists)
			from_month_num = MONTH_MAP.get(from_month)
			to_month_num = MONTH_MAP.get(to_month)

			# Start with filtering data based on year, month range, and village (just like list API)
			queryset = self.get_queryset().filter(
				date__year=year,
				date__month__gte=from_month_num,
				date__month__lte=to_month_num
			)

			village = request.query_params.get('village')
			if village:
				queryset = queryset.filter(villageName__icontains=village)

			# Apply the search term filtering to this already-filtered queryset
			queryset = queryset.filter(
				Q(patientName__icontains=search_term) |
				Q(villageName__icontains=search_term) |
				Q(date__icontains=search_term) |
				Q(age__icontains=search_term) |
				Q(signSymptoms__icontains=search_term) |
				Q(diagnosis__icontains=search_term) |
				Q(prescribedMedicine1__icontains=search_term) |
				Q(prescribedMedicine2__icontains=search_term) |
				Q(client_name__icontains=search_term) |
				Q(day__icontains=search_term) |
				Q(village__icontains=search_term) |
				Q(ageGroup__icontains=search_term) |
				Q(gender__icontains=search_term)
			)

			# Pagination logic (default limit is 50)
			start_index = int(request.query_params.get('start_index', 0))
			limit = 50  # Records per page
			queryset = queryset.order_by('pk')

			paginator = Paginator(queryset, limit)
			page_number = (start_index // limit) + 1

			try:
				page = paginator.page(page_number)
			except PageNotAnInteger:
				page = paginator.page(1)
			except EmptyPage:
				return Response({"message": "No more records available"}, status=status.HTTP_200_OK)

			# Serialize the paginated data
			paginated_data = self.serializer_class(page.object_list, many=True)

			# Create the response with pagination details
			api_response = {
				'status': 'success',
				'code': status.HTTP_200_OK,
				'total_records': paginator.count,
				'start_index': start_index,
				'limit': limit,
				'message': f'Search results for "{search_term}"',
				'patients': paginated_data.data,
			}

			return Response(api_response, status=status.HTTP_200_OK)

		except Exception as e:
			error_message = f'An error occurred while searching: {str(e)}'
			return Response({"status": "error", "message": error_message},
			                status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExcelReport(APIView):
	def get(self, request):
		data = Patientopdform.objects.all()  # Query data

		# Mapping of attribute names to display names
		headers_mapping = {
			'srNo': 'Sr.No.',
			'patientName': 'Patient Name',
			'date': 'Date',
			'village': 'Village',
			'villageName': 'Village Name',
			'category': 'N/F/SC/R',
			'gender': 'Gender',
			'age': 'Age',
			'day': 'Day',
			'month': 'Month',
			'ageGroup': 'Age Group',
			'week': 'Week',
			'mobileNo': 'Mobile No.',
			'signSymptoms': 'Sign and Symptoms',
			'physicalExamination': 'Physical Examination and Finding',
			'investigation': 'Investigation',
			'diagnosis': 'Diagnosis',
			'prescribedMedicine1': 'Prescribed medicine 1',
			'prescribedMedicine2': 'Prescribed medicine 2',
			'dosage': 'Dosage',
			'treatmentRemark': 'Treatment Remark',
			'client_name': 'Client Name'
		}

		# Initialize Excel workbook and sheet
		wb = openpyxl.Workbook()
		ws = wb.active

		# Write headers
		headers = list(headers_mapping.values())
		ws.append(headers)

		# Write data rows
		for row in data:
			row_data = [getattr(row, field) for field in headers_mapping.keys()]
			ws.append(row_data)

		# Get the current date
		current_date = datetime.date.today().strftime('%Y-%m-%d')
		thin_border = Border(left=Side(style='thin'),
		                     right=Side(style='thin'),
		                     top=Side(style='thin'),
		                     bottom=Side(style='thin'))

		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=24):
			for cell in row:
				cell.border = thin_border

		# Create HttpResponse object with Excel content type
		response = HttpResponse(
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = f'attachment; filename="OPD All Patient Records till {current_date}.xlsx"'

		# Save the Excel file to the HttpResponse
		wb.save(response)
		return response

class FilteredExcelReport(APIView):
	def get(self, request):

		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')
		if not year or not month or not village or not client_name:
			return Response({"error" : "Year, Month, Village and Client name are required"}, status=400)

		data=Patientopdform.objects.filter(year=year, month=month, village=village , client_name=client_name)

		# Mapping of attribute names to display names
		headers_mapping = {
			'srNo': 'Sr.No.',
			'patientName': 'Patient Name',
			'date': 'Date',
			'village': 'Village',
			'villageName': 'Village Name',
			'category': 'N/F/SC/R',
			'gender': 'Gender',
			'age': 'Age',
			'day': 'Day',
			'month': 'Month',
			'ageGroup': 'Age Group',
			'week': 'Week',
			'mobileNo': 'Mobile No.',
			'signSymptoms': 'Sign and Symptoms',
			'physicalExamination': 'Physical Examination and Finding',
			'investigation': 'Investigation',
			'diagnosis': 'Diagnosis',
			'prescribedMedicine1': 'Prescribed medicine 1',
			'prescribedMedicine2': 'Prescribed medicine 2',
			'dosage': 'Dosage',
			'treatmentRemark': 'Treatment Remark',
			'client_name': 'Client Name'
		}

		# Initialize Excel workbook and sheet
		wb = openpyxl.Workbook()
		ws = wb.active

		# Write headers
		headers = list(headers_mapping.values())
		ws.append(headers)

		# Write data rows
		for row in data:
			row_data = [getattr(row, field) for field in headers_mapping.keys()]
			ws.append(row_data)

		# Get the current date

		thin_border = Border(left=Side(style='thin'),
		                     right=Side(style='thin'),
		                     top=Side(style='thin'),
		                     bottom=Side(style='thin'))

		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=24):
			for cell in row:
				cell.border = thin_border

		# Create HttpResponse object with Excel content type
		response = HttpResponse(
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = f'attachment; filename="OPD Monthly Patient Records of{village}-{month}-{year}.xlsx"'

		# Save the Excel file to the HttpResponse
		wb.save(response)
		return response

# This WeeksExcelSheet class is dividing the whole excelsheet data of all months from Apr to August into Weeks - Now we are commenting it for future use
# class WeeksExcelSheet(APIView):
#     def get(self, request):
#         # Query data from Patientopdform model
#         data = Patientopdform.objects.all()  # Query data
#
#         # Mapping of attribute names to display names
#         headers_mapping = {
#             'srNo': 'Sr.No.',
#             'patientName': 'Patient Name',
#             'date': 'Date',
#             'villageName': 'Village Name',
#             'camp_name':'Camp Name',
#             'category': 'N/F/SC/R',
#             'gender': 'Gender',
#             'age': 'Age',
#             'day': 'Day',
#             'month': 'Month',
#             'ageGroup': 'Age Group',
#             'week': 'Week',
#             'mobileNo': 'Mobile No.',
#             'signSymptoms': 'Sign and Symptoms',
#             'physicalExamination': 'Physical Examination and Finding',
#             'investigation': 'Investigation',
#             'diagnosis': 'Diagnosis',
#             'prescribedMedicine1': 'Prescribed medicine 1',
#             'prescribedMedicine2': 'Prescribed medicine 2',
#             'dosage': 'Dosage',
#             'treatmentRemark': 'Treatment Remark',
#         }
#         media_path = settings.MEDIA_ROOT
#         if not os.path.exists(media_path):
#             os.makedirs(media_path)
#         # Initialize Excel workbook
#         wb = openpyxl.Workbook()
#
#         # Separate data for each week
#         weeks = data.values_list('week', flat=True).distinct()
#
#         for week in weeks:
#             ws = wb.create_sheet(title=f'Week {week}')
#             # Write headers
#             headers = list(headers_mapping.values())
#             ws.append(headers)
#
#             # Filter data by week
#             week_data = data.filter(week=week)
#
#             # Write data rows for that week
#             for row in week_data:
#                 row_data = [getattr(row, field) for field in headers_mapping.keys()]
#                 ws.append(row_data)
#
#         # Remove the default sheet created by openpyxl (if empty)
#         if 'Sheet' in wb.sheetnames:
#             default_sheet = wb['Sheet']
#             wb.remove(default_sheet)
#
#         # Get the current date
#         current_date = datetime.today().strftime('%Y-%m-%d')
#         file_path = os.path.join(media_path, f'Weeks_wise_report{current_date}.xlsx')
#         wb.save(file_path)
#         # Create HttpResponse object with Excel content type
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="Weeks_wise_report{current_date}.xlsx"'
#
#         # Save the Excel file to the HttpResponse
#         response.write(f'File saved successfully: {file_path}')
#         return response


# This WeeklyExcelReport class is generating the Report data from Week-wise-report(media folder)- Now we are commenting it for future use
# class WeeklyExcelReport(APIView):
#     def get(self, request):
#         # Query data from Patientopdform model
#         data = Patientopdform.objects.all()
#
#         # Mapping of attribute names to display names
#         headers_mapping = {
#             'srNo': 'Sr.No.',
#             'patientName': 'Patient Name',
#             'date': 'Date',
#             'villageName': 'Village Name',
#             'camp_name': 'Camp Name',
#             'category': 'N/F/SC/R',
#             'gender': 'Gender',
#              'age': 'Age',
#              'day': 'Day',
#              'month': 'Month',
#              'ageGroup': 'Age Group',
#              'week': 'Week',
#              'mobileNo': 'Mobile No.',
#              'signSymptoms': 'Sign and Symptoms',
#              'physicalExamination': 'Physical Examination and Finding',
#              'investigation': 'Investigation',
#              'diagnosis': 'Diagnosis',
#              'prescribedMedicine1': 'Prescribed medicine 1',
#              'prescribedMedicine2': 'Prescribed medicine 2',
#              'dosage': 'Dosage',
#              'treatmentRemark': 'Treatment Remark',
#         }
#
#         # Create a week-wise dictionary to hold dataframes
#         weeks_dataframes = {}
#
#         # Separate data for each week
#         weeks = data.values_list('week', flat=True).distinct()
#
#         # Create dataframe for each week
#         for week in weeks:
#             # Filter data by week
#             week_data = data.filter(week=week)
#
#             # Create dataframe for the current week
#             df = pd.DataFrame(list(week_data.values(*headers_mapping.keys())))
#             # Ensure column names are consistent
#             df.columns = [headers_mapping.get(col, col) for col in df.columns]
#             # Store dataframe in dictionary
#             weeks_dataframes[week] = df
#
#         # Initialize an Excel writer for output
#         output = io.BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # Process each week's dataframe
#             for week, df in weeks_dataframes.items():
#                 if df.empty:
#                     continue
#
#                 # Create a 'Count' column with value 1
#                 df['Count'] = 1
#                 # Ensure that 'Day' is in the DataFrame
#                 if 'Day' not in df.columns:
#                     print(f"Day column not found in Week {week}")
#                     continue
#                 # Create pivot table for the current week
#                 try:
#                     pivot_table = df.pivot_table(
#                         index='diagnosis',
#                         columns=['day', 'villageName', 'gender'],
#                         values='Count',
#                         aggfunc='sum',
#                         fill_value=0
#                     )
#                 except KeyError as e:
#                     # Skip this sheet if required columns are missing
#                     print(f"Skipping week {week} due to missing column: {e}")
#                     continue
#
#                 # Reset the index to make 'Diagnosis' the first column
#                 pivot_table.reset_index(inplace=True)
#
#                 # Write the pivot table to a new sheet in the output Excel file
#                 pivot_table.to_excel(writer, sheet_name=f'Weekly Report - Week {week}', index=True)
#
#         # Create HttpResponse object with Excel content type
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="Weekly_Reports_{datetime.today().strftime("%Y-%m-%d")}.xlsx"'
#
#         # Save the output to the HttpResponse
#         output.seek(0)
#         response.write(output.getvalue())
#
#         return response


class WeeklyReport(APIView):
	def get(self, request):
		# Sample data from your database
		data = Patientopdform.objects.all().values()
		# Convert query data to DataFrame
		df = pd.DataFrame(data)
		# Check if DataFrame is empty
		if df.empty:
			return Response({"error": "No data available"}, status=400)

		df['date'] = pd.to_datetime(df['date'], format='%d-%b-%y').dt.strftime('%Y-%m-%d')
		# Create a 'Count' column
		df['Count'] = 1

		# Initialize Excel writer for output
		output = io.BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# Step 1: Always create a default sheet with "No Data Available" message
			# df_empty = pd.DataFrame([['No Data Available']], columns=['Message'])
			# df_empty.to_excel(writer, sheet_name='No Data', index=False)

			# Step 2: Get unique values for 'week'
			weeks = df['week'].unique()
			sheet_added = False  # Flag to check if at least one valid sheet is added

			for week in weeks:
				# Filter data for the current week
				week_data = df[df['week'] == week]
				week_data = week_data.sort_values(by='day')
				# Check if there's data for the week
				if not week_data.empty:
					actual_days = week_data['day'].unique()
					# Create pivot table for the current week
					pivot_table = week_data.pivot_table(
						index=['diagnosis'],
						columns=['day', 'villageName', 'gender'],
						values='Count',
						aggfunc='sum',
						fill_value=0
					)

					# Add a 'Grand Total' column by summing across rows
					pivot_table['Grand Total For M/F'] = pivot_table.sum(axis=1)

					# Reset index
					pivot_table.reset_index(inplace=True)

					# Create dynamic multi-index for columns
					week_days = week_data[['day', 'date']].drop_duplicates().sort_values(
						by='day')
					day_date_map = dict(zip(week_days['day'], week_days['date']))

					# Generate dynamic multi-index tuples
					dynamic_columns = []
					for (day, village, gender) in pivot_table.columns[1:-1]:
						if day in actual_days:
							date = day_date_map.get(day, '')
							dynamic_columns.append(
								(day, date, village, gender))

					# Add 'Grand Total For M/F' at the end
					dynamic_columns.append(('Grand Total For M/F', '', '', ''))
					dynamic_columns = sorted(dynamic_columns,
					                         key=lambda x: pd.to_datetime(x[1]) if x[
						                                                               1] != '' else pd.NaT)
					dynamic_columns = [('Diagnosis', '', '', '')] + dynamic_columns

					# Assign the multi-index to the pivot table columns

					pivot_table.columns = pd.MultiIndex.from_tuples(dynamic_columns,
					                                                names=['Day', 'Date','Village','Gender'])

					# Write the pivot table to a new sheet
					pivot_table.to_excel(writer, sheet_name=f'Week {week}', index=True)
					sheet_added = True  # Mark that a valid sheet is added
					# Apply borders to the entire sheet
					workbook = writer.book
					worksheet = workbook[f'Week {week}']

					thin_border = Border(
						left=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin')
					)

					# Loop through all cells and apply border
					for row in worksheet.iter_rows():
						for cell in row:
							cell.border = thin_border
			# Step 6: Handle case where no sheets with actual data were added
			# if not sheet_added:
			#           df_empty.to_excel(writer, sheet_name='No Valid Data', index=False)

		# Set the pointer to the beginning of the output stream
		output.seek(0)

		# Return the Excel file as a FileResponse
		response = FileResponse(output, as_attachment=True, filename='WeeklyReport.xlsx')
		return response

class FilteredWeeklyReport(APIView):
	def get(self, request):
		# Sample data from your database
		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')
		if not year or not month or not village or not client_name:
			return Response({"error": "Year, Month, Village and Client name are required"}, status=400)

		data = Patientopdform.objects.filter(year=year, month=month, village=village, client_name=client_name).values()
		# Convert query data to DataFrame
		df = pd.DataFrame(data)
		# Check if DataFrame is empty
		if df.empty:
			return Response({"error": "No data available"}, status=400)

		# Create a 'Count' column
		df['Count'] = 1

		# Initialize Excel writer for output
		output = io.BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:

			weeks = df['week'].unique()
			sheet_added = False  # Flag to check if at least one valid sheet is added

			for week in weeks:
				# Filter data for the current week
				week_data = df[df['week'] == week]
				week_data = week_data.sort_values(by='day')
				# Check if there's data for the week
				if not week_data.empty:
					actual_days = week_data['day'].unique()
					# Create pivot table for the current week
					pivot_table = week_data.pivot_table(
						index=['diagnosis'],
						columns=['day', 'villageName', 'gender'],
						values='Count',
						aggfunc='sum',
						fill_value=0
					)

					# Add a 'Grand Total' column by summing across rows
					pivot_table['Grand Total For M/F'] = pivot_table.sum(axis=1)

					# Reset index
					pivot_table.reset_index(inplace=True)

					# Create dynamic multi-index for columns
					week_days = week_data[['day', 'date']].drop_duplicates().sort_values(
						by='day')
					day_date_map = dict(zip(week_days['day'], week_days['date']))

					# Generate dynamic multi-index tuples
					dynamic_columns = []
					for (day, village, gender) in pivot_table.columns[1:-1]:
						if day in actual_days:
							date = day_date_map.get(day, '')
							dynamic_columns.append(
								(day, date, village, gender))

					# Add 'Grand Total For M/F' at the end
					dynamic_columns.append(('Grand Total For M/F', '', '', ''))
					dynamic_columns = sorted(dynamic_columns,
					                         key=lambda x: pd.to_datetime(x[1]) if x[
						                                                               1] != '' else pd.NaT)
					dynamic_columns = [('Diagnosis', '', '', '')] + dynamic_columns

					# Assign the multi-index to the pivot table columns

					pivot_table.columns = pd.MultiIndex.from_tuples(dynamic_columns,
					                                                names=['Day', 'Date','Village','Gender'])

					# Write the pivot table to a new sheet
					pivot_table.to_excel(writer, sheet_name=f'Week {week}', index=True)
					sheet_added = True  # Mark that a valid sheet is added
					# Apply borders to the entire sheet
					workbook = writer.book
					worksheet = workbook[f'Week {week}']

					thin_border = Border(
						left=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin')
					)

					# Loop through all cells and apply border
					for row in worksheet.iter_rows():
						for cell in row:
							cell.border = thin_border

		# Set the pointer to the beginning of the output stream
		output.seek(0)

		# Return the Excel file as a FileResponse
		response = FileResponse(output, as_attachment=True, filename='WeeklyReport.xlsx')
		return response


# class MonthlyWeeklyReport(APIView):
#           def get(self, request):
#                     month = request.query_params.get('month')
#                     if not month:
#                               return Response({"error": "Month is required"}, status=400)
#
#                     try:
#                               month_datetime = pd.to_datetime(f'01-{month}', format='%d-%b-%Y')
#                     except ValueError:
#                               return Response({"error": "Invalid month format. Use 'MMM-YYYY' format, e.g. 'Jan-2024'"},
#                                               status=400)
#
#                     data = Patientopdform.objects.all().values()
#                     df = pd.DataFrame(data)
#                     if df.empty:
#                               return Response({"error": f"No data available for the month:{month}"}, status=400)
#                     df['date'] = pd.to_datetime(df['date'], format='%d-%b-%y')
#                     # Filter by month and year
#                     df = df[(df['date'].dt.month == month_datetime.month) & (df['date'].dt.year == month_datetime.year)]
#
#                     df['Count'] = 1
#
#                     # Initialize Excel writer for output
#                     output = io.BytesIO()
#                     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                               # Step 2: Get unique values for 'week'
#                               weeks = df['week'].unique()
#                               for week in weeks:
#                                         # Filter data for the current week
#                                         week_data = df[df['week'] == week]
#                                         week_data = week_data.sort_values(by='day')
#                                         # Check if there's data for the week
#                                         if not week_data.empty:
#                                                   actual_days = week_data['day'].unique()
#                                                   # Create pivot table for the current week
#                                                   pivot_table = week_data.pivot_table(
#                                                             index=['diagnosis'],
#                                                             columns=['day', 'villageName', 'gender'],
#                                                             values='Count',
#                                                             aggfunc='sum',
#                                                             fill_value=0
#                                                   )
#
#                                                   # Add a 'Grand Total' column by summing across rows
#                                                   pivot_table['Grand Total For M/F'] = pivot_table.sum(axis=1)
#
#                                                   # Reset index
#                                                   pivot_table.reset_index(inplace=True)
#
#                                                   # Create dynamic multi-index for columns
#                                                   week_days = week_data[['day', 'date']].drop_duplicates().sort_values(
#                                                             by='day')
#                                                   day_date_map = dict(zip(week_days['day'],
#                                                                           week_days['date'].dt.strftime('%Y-%m-%d')))
#
#                                                   # Generate dynamic multi-index tuples
#                                                   dynamic_columns = []
#                                                   for (day, village, gender) in pivot_table.columns[1:-1]:
#                                                             if day in actual_days:
#                                                                       date = day_date_map.get(day, '')
#                                                                       dynamic_columns.append(
#                                                                                 (day, date, village, gender))
#
#                                                   # Add 'Grand Total For M/F' at the end
#                                                   dynamic_columns.append(('Grand Total For M/F', '', '', ''))
#                                                   dynamic_columns = sorted(dynamic_columns,
#                                                                            key=lambda x: pd.to_datetime(x[1]) if x[
#                                                                                                                            1] != '' else pd.NaT)
#                                                   dynamic_columns = [('Diagnosis', '', '', '')] + dynamic_columns
#
#                                                   # Assign the multi-index to the pivot table columns
#                                                   pivot_table.columns = pd.MultiIndex.from_tuples(dynamic_columns,
#                                                                                                   names=['Day', 'Date',
#                                                                                                          'Village',
#                                                                                                          'Gender'])
#
#                                                   # Write the pivot table to a new sheet
#                                                   pivot_table.to_excel(writer, sheet_name=f'Week {week}', index=True)
#                                                   sheet_added = True  # Mark that a valid sheet is added
#
#                                                   # Apply borders to the entire sheet
#                                                   workbook = writer.book
#                                                   worksheet = workbook[f'Week {week}']
#
#                                                   thin_border = Border(
#                                                             left=Side(style='thin'),
#                                                             right=Side(style='thin'),
#                                                             top=Side(style='thin'),
#                                                             bottom=Side(style='thin')
#                                                   )
#
#                                                   # Loop through all cells and apply border
#                                                   for row in worksheet.iter_rows():
#                                                             for cell in row:
#                                                                       cell.border = thin_border
#
#                     # Set the pointer to the beginning of the output stream
#                     output.seek(0)
#
#                     # Return the Excel file as a FileResponse
#                     response = FileResponse(output, as_attachment=True,
#                                             filename=f'Monthly-Week-Wise-Report_{month}.xlsx')
#                     return response

class VillageWiseGenderReport(APIView):
	def get(self, request):
		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')
		header_text = f"{village} Village Wise Gender Wise Report for Month {month}-{year}"
		# Ensure that year, month and village are provided
		if not year:
			return Response({"error": "Year is required"}, status=400)
		if not month:
			return Response({"error": "Month is required"}, status=400)
		if not village:
			return Response({"error": "Village is required"}, status=400)
		if not client_name:
			return Response({"error": "Client Name is required"}, status=400)

		# Fetch data from the database, filtered by village, month, and year
		data = Patientopdform.objects.filter(year=year, month=month, village=village, client_name=client_name).values(
			'villageName',
			'gender')

		if not data:
			return Response({"error": f"No data available for the month: {month}"}, status=400)

		# Create a DataFrame from the filtered data
		df = pd.DataFrame(data)

		# Create pivot table for village-wise gender counts
		pivot_table = df.pivot_table(
			index='villageName',
			columns='gender',
			aggfunc='size',  # Count occurrences for each gender
			fill_value=0  # Replace NaN with 0 for missing genders
		)

		# Add a 'Total' column for the sum of all genders
		pivot_table['Total'] = pivot_table.sum(axis=1)

		# Ensure the table includes columns for 'Male', 'Female', and 'Other' even if missing in the data
		for col in ['Male', 'Female', 'Other']:
			if col not in pivot_table.columns:
				pivot_table[col] = 0

		# Reorder columns to match the required report structure
		pivot_table = pivot_table[['Male', 'Female', 'Other', 'Total']]

		# Initialize Excel writer for output
		output = io.BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# Write the village-wise gender report to an Excel sheet
			pivot_table.to_excel(writer, sheet_name=f'VillageWiseReport_{month}', index=True)

			# Apply borders to the entire sheet
			workbook = writer.book
			worksheet = workbook[f'VillageWiseReport_{month}']
			# worksheet["A1"] = header_text
			# worksheet["A1"].font = Font(bold=True , size=16)

			thin_border = Border(
				left=Side(style='thin'),
				right=Side(style='thin'),
				top=Side(style='thin'),
				bottom=Side(style='thin')
			)

			# Loop through all cells and apply border
			for row in worksheet.iter_rows():
				for cell in row:
					cell.border = thin_border

		# Set the pointer to the beginning of the output stream
		output.seek(0)

		# Return the Excel file as a FileResponse
		response = FileResponse(output, as_attachment=True, filename=f'VillageWiseGenderReport_{month}.xlsx')
		return response


class VillageWiseAgeGroupReport(APIView):
	def get(self, request):
		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')

		# Ensure that year, month and village are provided
		if not year:
			return Response({"error": "Year is required"}, status=400)
		if not month:
			return Response({"error": "Month is required"}, status=400)
		if not village:
			return Response({"error": "Village is required"}, status=400)
		if not client_name:
			return Response({"error": "Client Name is required"}, status=400)

		# Fetch data from the database, filtered by village, month, and year
		data = Patientopdform.objects.filter(year=year, month=month, village=village, client_name=client_name).values(
			'villageName',
			'ageGroup',
			'date')

		# Convert the data to a pandas DataFrame
		df = pd.DataFrame(data)

		# Check if DataFrame is empty (i.e., no data available)
		if df.empty:
			return Response(
				{"error": f"No data available for the month: {month} in village: {village}"}, status=400)

		# Create pivot table for village-wise age group counts
		pivot_table = df.pivot_table(
			index='villageName',
			columns='ageGroup',
			aggfunc='size',  # Count occurrences for each age group
			fill_value=0  # Replace NaN with 0 for missing age groups
		)

		# Add a 'Total' column for the sum of all age groups
		pivot_table['Total'] = pivot_table.sum(axis=1)

		# Dynamically get the distinct age groups from the data
		age_labels = df['ageGroup'].unique().tolist()

		# Ensure the table includes all age groups even if missing in the data
		for col in age_labels:
			if col not in pivot_table.columns:
				pivot_table[col] = 0

		# Reorder columns to match the required report structure
		pivot_table = pivot_table[age_labels + ['Total']]

		# Initialize Excel writer for output
		output = io.BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# Write the village-wise age group report to an Excel sheet
			pivot_table.to_excel(writer, sheet_name=f'VillageWiseAgeGroup_{month}', index=True)

			# Apply borders to the entire sheet
			workbook = writer.book
			worksheet = workbook[f'VillageWiseAgeGroup_{month}']

			thin_border = Border(
				left=Side(style='thin'),
				right=Side(style='thin'),
				top=Side(style='thin'),
				bottom=Side(style='thin')
			)

			# Loop through all cells and apply border
			for row in worksheet.iter_rows():
				for cell in row:
					cell.border = thin_border

		# Set the pointer to the beginning of the output stream
		output.seek(0)

		# Return the Excel file as a FileResponse
		response = FileResponse(output, as_attachment=True, filename=f'VillageWiseAgeGroupReport_{month}.xlsx')
		return response


class MonthlySummaryReport(APIView):
	def get(self, request):
		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')
		if not year:
			return Response({"error": "Year is required"}, status=400)
		if not month:
			return Response({"error": "Month is required"}, status=400)
		if not village:
			return Response({"error": "Village is required"}, status=400)
		if not client_name:
			return Response({"error": "Client Name is required"}, status=400)
			# Fetch data from the database (filtered by year, month, and village)
		data = Patientopdform.objects.filter(client_name=client_name, village=village, year=year, month=month).values()

		df = pd.DataFrame(data)
		if df.empty:
			return Response({
				"error": f"No data available for Year: {year}, Month: {month}, Village: {village}"},
				status=400)

		df['Count'] = 1
		# Generate summary report
		output = io.BytesIO()

		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# Step 1: Get unique village names dynamically
			village_names = df['villageName'].unique()

			# Step 2: Create a pivot table summarizing data by disease and village
			pivot_table = df.pivot_table(
				index='diagnosis',
				columns='villageName',
				values='Count',
				aggfunc='sum',
				fill_value=0
			)

			# Step 3: Add 'Sr No' column at the beginning
			pivot_table.reset_index(inplace=True)
			pivot_table.insert(0, 'Sr No', range(1, len(pivot_table) + 1))
			pivot_table['Total'] = pivot_table[village_names].sum(axis=1)

			# Step 4: Add a 'Total' row (village-wise totals, vertical sum)
			total_row = pivot_table[village_names].sum(axis=0).to_frame().T
			total_row['diagnosis'] = 'Total'
			total_row['Sr No'] = ''  # No Sr No for total row
			total_row['Total'] = total_row[village_names].sum(axis=1)

			# Append the 'Total' row to the pivot table
			pivot_table = pd.concat([pivot_table, total_row], ignore_index=True)

			# Step 5: Write the pivot table to Excel
			pivot_table.to_excel(writer, sheet_name=f'Monthly Summary {month}-{year}', index=False)
			# Apply borders to the entire sheet
			workbook = writer.book
			worksheet = workbook[f'Monthly Summary {month}-{year}']
			# Add header "Summary Village Wise Disease Wise Count"
			worksheet.insert_rows(1)
			worksheet.merge_cells(start_row=1, start_column=1, end_row=1,
			                      end_column=len(pivot_table.columns))
			header_cell = worksheet.cell(row=1, column=1)
			header_cell.value = "Summary Village Wise Disease Wise Count"
			header_cell.font = Font(size=14, bold=True)
			header_cell.alignment = Alignment(horizontal='center')
			thin_border = Border(
				left=Side(style='thin'),
				right=Side(style='thin'),
				top=Side(style='thin'),
				bottom=Side(style='thin')
			)

			for row in worksheet.iter_rows():
				for cell in row:
					cell.border = thin_border
				# Step 6: Add 'Total' row at the bottom
			total_row = worksheet.max_row

			for col_idx in range(3, len(village_names) + 3):
				# Sum village-wise totals
				cell_value = f"=SUM({worksheet.cell(row=3, column=col_idx).coordinate}:{worksheet.cell(row=total_row - 1, column=col_idx).coordinate})"
				worksheet.cell(row=total_row, column=col_idx, value=cell_value)

			grand_total = f"=SUM({worksheet.cell(row=3, column=len(pivot_table.columns)).coordinate}:{worksheet.cell(row=total_row - 1, column=len(pivot_table.columns)).coordinate})"
			# Sum the 'Total' column
			worksheet.cell(row=total_row, column=len(pivot_table.columns), value=grand_total)
			# value=f"=SUM({worksheet.cell(row=3, column=len(pivot_table.columns)).coordinate}:{worksheet.cell(row=total_row - 1, column=len(pivot_table.columns)).coordinate})")

			font_style = Font(bold=True)
			# Step 7: Add 'Percentage' column (for each disease)
			worksheet.cell(row=2, column=len(pivot_table.columns) + 1, value="Percentage").font = font_style
			grand_total_cell = worksheet.cell(row=total_row, column=len(pivot_table.columns)).coordinate

			for row_idx in range(3, total_row):
				total_value_cell = worksheet.cell(row=row_idx, column=len(pivot_table.columns)).coordinate
				percentage_formula = f"=ROUND(({total_value_cell}/{grand_total_cell})*100, 0)"
				cell = worksheet.cell(row=row_idx, column=len(pivot_table.columns) + 1, value=percentage_formula)
				cell.border = thin_border
				# cell.font=font_style
				# Apply border to 'Percentage' column as well

			for row in range(2, total_row):
				cell = worksheet.cell(row=row, column=len(pivot_table.columns) + 1)
				cell.border = thin_border

			# Set the pointer to the beginning of the output stream
		output.seek(0)

		# Return the Excel file as a FileResponse
		response = FileResponse(output, as_attachment=True, filename=f'Monthly-Summary-Report_{month}-{year}.xlsx')
		return response


class SummaryDiseaseWiseWeeklyReport(APIView):
	def get(self, request):
		year = request.query_params.get('year')
		month = request.query_params.get('month')
		village = request.query_params.get('village')
		client_name = request.query_params.get('client_name')
		if not client_name:
			return Response({"error": "Client Name is required"}, status=400)
		if not year:
			return Response({"error": "Year is required"}, status=400)
		if not month:
			return Response({"error": "Month is required"}, status=400)
		if not village:
			return Response({"error": "Village is required"}, status=400)

		# Fetch data from the database (filtered by year, month, and village)
		data = Patientopdform.objects.filter(
			client_name=client_name,
			village=village,
			year=year,
			month=month
		).values()

		df = pd.DataFrame(data)
		if df.empty:
			return Response({
				"error": f"No data available for Year: {year}, Month: {month}, Village: {village}"},
				status=400)

		df['Count'] = 1

		# Pivot table for weekly data, disease-wise and gender-wise
		pivot_table = df.pivot_table(
			index='diagnosis',
			columns=['week', 'gender'],
			values='Count',
			aggfunc='sum',
			fill_value=0
		)

		# Add 'Sr No' column at the beginning
		pivot_table.reset_index(inplace=True)
		pivot_table.insert(0, 'Sr No', range(1, len(pivot_table) + 1))

		# Calculate Male/Female totals for each diagnosis
		male_columns = [col for col in pivot_table.columns if 'Male' in col]
		female_columns = [col for col in pivot_table.columns if 'Female' in col]

		pivot_table['Total Male'] = pivot_table[male_columns].sum(axis=1)
		pivot_table['Total Female'] = pivot_table[female_columns].sum(axis=1)
		pivot_table['Grand Total'] = pivot_table['Total Male'] + pivot_table['Total Female']

		# Total counts for all weeks combined (vertical sum)
		total_row = pivot_table[male_columns + female_columns].sum().to_frame().T
		total_row['diagnosis'] = 'Total'
		total_row['Sr No'] = ''
		total_row['Total Male'] = total_row[male_columns].sum(axis=1)
		total_row['Total Female'] = total_row[female_columns].sum(axis=1)
		total_row['Grand Total'] = total_row['Total Male'] + total_row['Total Female']

		# Append total row to pivot table
		pivot_table = pd.concat([pivot_table, total_row], ignore_index=True)
		# Calculate percentage based on the 'Grand Total'
		grand_total_sum = pivot_table['Grand Total'].iloc[:-1].sum()  # Exclude the total row itself
		print("GRand Total sum", grand_total_sum)
		pivot_table['Percentage'] = (pivot_table['Grand Total'] / grand_total_sum * 100).round(2)

		# Writing to Excel
		output = io.BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# Write pivot table to Excel sheet
			pivot_table.to_excel(writer,
			                     sheet_name=f'Summary-DiseaseWise-WeekWise-PatientCount-Report-{month}-{year}',
			                     index=True)

			# Get the Excel workbook and worksheet objects
			workbook = writer.book
			worksheet = workbook[f'Summary-DiseaseWise-WeekWise-PatientCount-Report-{month}-{year}']

			# Apply formatting: borders, font size, alignment, etc.
			thin_border = Border(
				left=Side(style='thin'),
				right=Side(style='thin'),
				top=Side(style='thin'),
				bottom=Side(style='thin')
			)

			# Apply borders to all cells
			for row in worksheet.iter_rows():
				for cell in row:
					cell.border = thin_border

			# Add "Grand Total" row sum
			total_row_idx = worksheet.max_row
			for col_idx in range(3, len(pivot_table.columns) - 1):
				# Sum formula for Grand Total columns
				sum_formula = f"=SUM({worksheet.cell(row=3, column=col_idx).coordinate}:{worksheet.cell(row=total_row_idx - 1, column=col_idx).coordinate})"
				worksheet.cell(row=total_row_idx, column=col_idx, value=sum_formula)
				# Format percentage column
			percentage_column_idx = len(pivot_table.columns)
			for row_idx in range(2, total_row_idx):  # Start from row 2 (excluding header)
				worksheet.cell(row=row_idx, column=percentage_column_idx)

		output.seek(0)
		response = FileResponse(output, as_attachment=True,
		                        filename=f'Summary-DiseaseWise-WeekWise-PatientCount-Report-{month}-{year}.xlsx')
		return response


class PatientHistoryAPI(APIView):
	def get(self, request, format=None):
		# Get the patientName from query parameters
		patient_name = request.query_params.get('patientName', None)

		# If patientName is not provided, return a bad request response
		if not patient_name:
			return Response({
				"status": "error",
				"code": 400,
				"message": "The 'patientName' parameter is required.",
				"data": None
			}, status=status.HTTP_400_BAD_REQUEST)

		try:
			# Fetch the latest record for the given patientName ordered by date (descending)
			latest_record = Patientopdform.objects.filter(patientName=patient_name).order_by(
				'-date').first()

			if latest_record:
				# Serialize only the investigation and diagnosis fields
				serializer = PatientHistorySerializer(latest_record)
				return Response({
					"status": "success",
					"code": 200,
					"message": "Latest patient history retrieved successfully.",
					"data": serializer.data
				}, status=status.HTTP_200_OK)
			else:
				# If no record is found for the patientName, return a not found response
				return Response({
					"status": "error",
					"code": 404,
					"message": f"No history found for the patient name: {patient_name}."
				}, status=status.HTTP_404_NOT_FOUND)

		except Exception as e:
			# In case of any server error, return a server error response
			return Response({
				"status": "error",
				"code": 500,
				"message": "An error occurred while retrieving the patient's history.",
				"data": str(e)
			}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
