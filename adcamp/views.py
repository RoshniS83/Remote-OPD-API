from datetime import datetime
import datetime
import openpyxl
import pandas as pd
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from .models import ADCamp
from .serializers import ADCampSerializer
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Count

class ADCampAPI(ModelViewSet):
    queryset = ADCamp.objects.all()
    serializer_class = ADCampSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            api_response = {
                'status': 'success',
                'code': status.HTTP_201_CREATED,
                'message': 'ADCamp record added successfully',
                'new_record': serializer.data,
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to add ADCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.serializer_class(instance)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'adcamp': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve ADCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_404_NOT_FOUND,
                'message': error_message
            }
            return Response(error_response)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.serializer_class(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'ADCamp record updated successfully',
                'updated_record': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to update ADCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def partial_update(self, request, *args, **kwargs):
        try:
            kwargs['partial'] = True
            response = self.update(request, *args, **kwargs)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'ADCamp record updated successfully',
                'updated_record': response.data
            }
            return Response(api_response, status=status.HTTP_200_OK)
        except Exception as e:
            error_message = 'Failed to partially update ADCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            instance.delete()
            api_response = {
                'status': 'success',
                'code': status.HTTP_204_NO_CONTENT,
                'message': 'ADCamp record deleted successfully'
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to delete ADCamp record: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.serializer_class(queryset, many=True)
            api_response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'adcamps': serializer.data
            }
            return Response(api_response)
        except Exception as e:
            error_message = 'Failed to retrieve ADCamp records: {}'.format(str(e))
            error_response = {
                'status': 'error',
                'code': status.HTTP_400_BAD_REQUEST,
                'message': error_message
            }
            return Response(error_response)


class ADCampExcelSheet(APIView):
    def get(self, request):
        data = ADCamp.objects.all()
        headers_mapping = {
            'SrNo' :' Sr No',
            'village' : 'Main Village',
            'date': 'Date',
            'year': 'Year',
            'month': 'Month',
            'name' : 'Name',
            'age' : 'Age',
            'contact':'Contact',
            'standard': 'Standard',
            'weight': 'Weight',
            'height': 'Height',
            'villageName' : 'Village Name',
            'HB' : 'HB',
            'HBReadings' : 'HB Readings',
            'BMI': 'BMI',
            'BMIReadings':'BMI Readings',
            'client_name' :'Client Name'
        }
        wb = openpyxl.Workbook()
        ws=wb.active
        ws.merge_cells('A1:P1')
        ws['A1'] = 'Arogaya Dhansampadha Camp'
        ws['A1'].font = Font(bold= True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers= list(headers_mapping.values())
        ws.append(headers)

        for cell in ws[2]:
            cell.font = Font(bold=True)
        for row in data:
            row_data=[getattr(row,field) for field in headers_mapping.keys()]
            ws.append(row_data)
            # Bold the header row

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = thin_border

        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="AD CAMP Records {current_date}.xlsx"'
        wb.save(response)
        return response


class ADCampHBMonthlyReport(APIView):
    def get(self, request):
        # Query data from the ADCamp model
        # data = ADCamp.objects.values('subvillage', 'gender', 'ADReadings').annotate(count=Count('SrNo'))
        # Get query parameters for filtering
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        village = request.query_params.get('village')
        client_name = request.query_params.get('client_name')

        # Validate required parameters
        if not year or not month or not village or not client_name:
            return Response({"error": "Year, Month, Village and Client Name are required"}, status=400)
            # Filter data from ADCamp model based on the provided parameters
        data = ADCamp.objects.filter(client_name=client_name, village=village, year=year, month=month).values('villageName', 'HBReadings').annotate(count=Count('SrNo'))

        # Convert to pandas dataframe
        df = pd.DataFrame(data)

        # Check if the DataFrame is empty
        if df.empty:
            return Response({"error": f"No data available for year: {year}, Month: {month}, Village: {village}"},
                            status=400)
            # Create a pivot table with 'subvillage' and 'gender' as index and 'ADReadings' as columns
        pivot_table = pd.pivot_table(df,
                                     index=['villageName'],
                                     columns='HBReadings',
                                     values='count',
                                     aggfunc='sum',
                                     fill_value=0)

        # Add Total column for each row
        pivot_table['Total'] = pivot_table.sum(axis=1)

        # Add Grand Total for each village by summing across genders
        # grand_total = pivot_table.groupby('villageName')['Total'].sum()
        # pivot_table['Grand Total Village'] = pivot_table.index.get_level_values('villageName').map(grand_total)

        # Reorder the columns to match the required order
        required_columns = ['Mild Anemia', 'Moderate Anemia', 'Severe Anemia', 'Healthy', 'Total']
        for col in required_columns:
            if col not in pivot_table.columns:
                pivot_table[col] = 0  # Add missing columns if necessary

        pivot_table = pivot_table[required_columns]
        #compute the total row across all villages
        total_row= pivot_table.sum(axis=0)
        total_row.name = 'Total'
        # Append the Total row to the pivot table using pd.concat()
        pivot_table = pd.concat([pivot_table, pd.DataFrame(total_row).T])
        # Convert pivot table to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AD Camp Report'

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Village Wise Aarogya Dhansampada Camp HB Screening Report {month}-{year}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Sr. No', 'Village Name'] + required_columns
        ws.append(headers)

        # Bold the header row
        for cell in ws[2]:
            cell.font = Font(bold=True)

        # Add data to the Excel sheet
        sr_no = 1
        current_row = 3
        for village, row_data in pivot_table.iterrows():
            if village != 'Total':
                ws.cell(row=current_row, column=1).value = sr_no  # Sr. No
                sr_no+=1
            else:
                ws.cell(row=current_row, column=1).value = ''
            ws.cell(row=current_row, column=2).value = village  # Village Name

            # Fill the rest of the values (Mild, Moderate, etc.)
            for col_idx, value in enumerate(row_data.tolist(), start=3):
                ws.cell(row=current_row, column=col_idx).value = value  # Assign values

            current_row += 1


            # Apply thin border to all cells
        thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border

            # Prepare the response to download the file
        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="AD_CAMP_HB_Report_{current_date}.xlsx"'
        wb.save(response)
        return response


class ADCampBMIMonthlyReport(APIView):
    def get(self, request):
        # Query data from the ADCamp model
        # data = ADCamp.objects.values('subvillage', 'gender', 'ADReadings').annotate(count=Count('SrNo'))
        # Get query parameters for filtering
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        village = request.query_params.get('village')
        client_name = request.query_params.get('client_name')

        # Validate required parameters
        if not year or not month or not village or not client_name:
            return Response({"error": "Year, Month, Village and Client Name are required"}, status=400)
            # Filter data from ADCamp model based on the provided parameters
        data = ADCamp.objects.filter(client_name=client_name, village=village, year=year, month=month).values('villageName', 'BMIReadings').annotate(count=Count('SrNo'))

        # Convert to pandas dataframe
        df = pd.DataFrame(data)

        # Check if the DataFrame is empty
        if df.empty:
            return Response({"error": f"No data available for year: {year}, Month: {month}, Village: {village}"},
                            status=400)
            # Create a pivot table with 'subvillage' and 'gender' as index and 'ADReadings' as columns
        pivot_table = pd.pivot_table(df,
                                     index=['villageName'],
                                     columns='BMIReadings',
                                     values='count',
                                     aggfunc='sum',
                                     fill_value=0)

        # Add Total column for each row
        pivot_table['Total'] = pivot_table.sum(axis=1)

        # Add Grand Total for each village by summing across genders
        # grand_total = pivot_table.groupby('villageName')['Total'].sum()
        # pivot_table['Grand Total Village'] = pivot_table.index.get_level_values('villageName').map(grand_total)

        # Reorder the columns to match the required order
        required_columns = ['Underweight', 'Healthy Weight', 'Overweight', 'Obese', 'Severely Obese','Morbidly Obese','Total']
        for col in required_columns:
            if col not in pivot_table.columns:
                pivot_table[col] = 0  # Add missing columns if necessary

        pivot_table = pivot_table[required_columns]
        #compute the total row across all villages
        total_row= pivot_table.sum(axis=0)
        total_row.name = 'Total'
        # Append the Total row to the pivot table using pd.concat()
        pivot_table = pd.concat([pivot_table, pd.DataFrame(total_row).T])
        # Convert pivot table to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AD Camp Report'

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = 'Village Wise Aarogya Dhansampada Camp BMI Report {month}-{year}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Sr. No', 'Village Name'] + required_columns
        ws.append(headers)

        # Bold the header row
        for cell in ws[2]:
            cell.font = Font(bold=True)

        # Add data to the Excel sheet
        sr_no = 1
        current_row = 3
        for village, row_data in pivot_table.iterrows():
            if village != 'Total':
                ws.cell(row=current_row, column=1).value = sr_no  # Sr. No
                sr_no+=1
            else:
                ws.cell(row=current_row, column=1).value = ''
            ws.cell(row=current_row, column=2).value = village  # Village Name

            # Fill the rest of the values (Mild, Moderate, etc.)
            for col_idx, value in enumerate(row_data.tolist(), start=3):
                ws.cell(row=current_row, column=col_idx).value = value  # Assign values

            current_row += 1


            # Apply thin border to all cells
        thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border

            # Prepare the response to download the file
        current_date = datetime.date.today().strftime('%Y-%m-%d')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="AD_CAMP_BMI_Report_{current_date}.xlsx"'
        wb.save(response)
        return response
